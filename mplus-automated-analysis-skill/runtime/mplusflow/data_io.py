from __future__ import annotations

import json
import math
import os
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from .utils import ensure_dir, safe_copy, sha256_file, write_json
from .sample_size import advisory_markdown, sample_size_advisory


SENTINEL_CANDIDATES = {-999999, -99999, -9999, -999, -99, 99, 999, 9999, 99999, 999999}
INTERNAL_MISSING = -999999.0


@dataclass
class PreparedLPAData:
    project_dir: Path
    runtime_dir: Path
    audit_runtime_dir: Path
    data_file: Path
    original_copy: Path
    spec_file: Path
    audit_file: Path
    variable_map_file: Path
    variable_map_csv: Path
    spec: dict[str, Any]
    variable_map: pd.DataFrame
    id_map: pd.DataFrame
    analysis_df: pd.DataFrame
    expected_mplus_n: int


def _read_text_table(
    path: Path,
    text_columns: list[str] | None = None,
    preserve_text_columns: list[str] | None = None,
) -> tuple[pd.DataFrame, str]:
    encodings = ["utf-8-sig", "utf-8", "gb18030", "cp936", "latin1"]
    last_exc: Exception | None = None
    for enc in encodings:
        try:
            if text_columns:
                preserve_indices = {
                    index: "string"
                    for index, name in enumerate(text_columns)
                    if name in (preserve_text_columns or [])
                }
                frame = pd.read_csv(
                    path, sep=None, engine="python", header=None, encoding=enc,
                    dtype=preserve_indices,
                )
                if frame.shape[1] != len(text_columns):
                    raise ValueError(
                        f"无表头文本实际有 {frame.shape[1]} 列，但提供了 {len(text_columns)} 个变量名。"
                        "已停止读取，防止首列被当成索引或全部变量错位。"
                    )
                frame.columns = text_columns
                return frame, enc
            return pd.read_csv(
                path, sep=None, engine="python", encoding=enc,
                dtype={name: "string" for name in (preserve_text_columns or [])},
            ), enc
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
    raise ValueError(f"无法读取文本数据：{path.name}。请确认分隔符、编码和是否有表头。原始错误：{last_exc}")


def load_dataframe(
    path: Path,
    text_columns: list[str] | None = None,
    preserve_text_columns: list[str] | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    ext = path.suffix.lower()
    meta: dict[str, Any] = {"格式": ext, "来源文件": str(path)}
    if ext in {".xlsx", ".xls"}:
        df = pd.read_excel(path, dtype={name: "string" for name in (preserve_text_columns or [])})
        if ext == ".xlsx" and preserve_text_columns:
            from openpyxl import load_workbook

            ws = load_workbook(path, read_only=True, data_only=True).active
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for name in preserve_text_columns:
                if name not in headers:
                    continue
                column_index = headers.index(name) + 1
                for (cell,) in ws.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                    if cell.data_type == "n" and isinstance(cell.value, (int, float)) and abs(cell.value) >= 10**15:
                        raise ValueError(
                            f"Excel 中 ID 变量“{name}”含超过15位的数值单元格，Excel 可能已改变其精度。"
                            "请把该列在源文件中改为文本后重新导入。"
                        )
    elif ext == ".csv":
        df, text_encoding = _read_text_table(path, preserve_text_columns=preserve_text_columns)
        meta["文本编码"] = text_encoding
    elif ext in {".txt", ".dat", ".tsv"}:
        df, text_encoding = _read_text_table(path, text_columns=text_columns, preserve_text_columns=preserve_text_columns)
        meta["文本编码"] = text_encoding
    elif ext in {".sav", ".zsav", ".dta"}:
        try:
            import pyreadstat  # type: ignore
        except ImportError as exc:
            raise RuntimeError("当前开发环境缺少 pyreadstat。正式发布的独立 Runtime 会内置该依赖；源码运行请先安装 pyreadstat。") from exc
        if ext in {".sav", ".zsav"}:
            df, rs_meta = pyreadstat.read_sav(str(path), apply_value_formats=False)
        else:
            df, rs_meta = pyreadstat.read_dta(str(path), apply_value_formats=False)
        meta["统计软件元数据"] = {
            "列数": getattr(rs_meta, "number_columns", None),
            "行数": getattr(rs_meta, "number_rows", None),
            "变量标签": {
                str(name): str(label) if label is not None else ""
                for name, label in zip(
                    getattr(rs_meta, "column_names", []) or [],
                    getattr(rs_meta, "column_labels", []) or [],
                )
            },
            "用户缺失范围": getattr(rs_meta, "missing_ranges", {}) or {},
            "用户缺失离散值": getattr(rs_meta, "missing_user_values", {}) or {},
        }
        for name in preserve_text_columns or []:
            if name in df.columns and pd.api.types.is_numeric_dtype(df[name]):
                numeric_id = pd.to_numeric(df[name], errors="coerce").abs()
                if (numeric_id >= 10**15).any():
                    raise ValueError(
                        f"统计软件文件中的 ID 变量“{name}”以数值存储且超过15位，原格式可能无法保证精度。"
                        "请在源软件中将 ID 转为字符串后再导出。"
                    )
    else:
        raise ValueError(f"暂不支持的数据格式：{ext}。支持 SAV/Z SAV、XLSX/XLS、CSV、DTA、TXT/DAT/TSV。")

    # 保留列名原貌，但删除首尾空格；重复列名属于重大风险。
    df.columns = [str(c).strip() for c in df.columns]
    duplicates = pd.Index(df.columns)[pd.Index(df.columns).duplicated()].tolist()
    if duplicates:
        raise ValueError(f"数据中存在重复列名：{duplicates}。请先处理，否则无法建立可靠变量映射。")
    meta["行数"] = int(len(df))
    meta["列数"] = int(df.shape[1])
    return df, meta


def write_verified_mplus_data(
    frame: pd.DataFrame,
    destinations: list[Path],
    missing_value: float = INTERNAL_MISSING,
) -> dict[str, Any]:
    """Write numeric free-format data and verify the exact Mplus-facing structure."""
    if frame.empty or frame.shape[1] < 2:
        raise ValueError("Mplus 分析数据必须至少包含一行和两个字段（ROWID 与分析变量）。")
    expected = frame.astype(float).fillna(float(missing_value)).to_numpy()
    if not np.isfinite(expected).all():
        raise ValueError("Mplus 分析数据含 NaN 以外的无穷值，已停止导出。")
    result: dict[str, Any] | None = None
    for destination in destinations:
        destination.parent.mkdir(parents=True, exist_ok=True)
        frame.to_csv(
            destination,
            sep=" ",
            header=False,
            index=False,
            na_rep=str(int(missing_value)),
            # 17 significant digits preserve an IEEE-754 double on text round-trip.
            float_format="%.17g",
            encoding="ascii",
            lineterminator="\n",
        )
        raw = destination.read_bytes()
        if not raw.isascii():
            raise RuntimeError(f"Mplus 数据文件不是纯 ASCII：{destination.name}")
        lines = [line for line in raw.splitlines() if line.strip()]
        widths = [len(line.split()) for line in lines]
        if len(lines) != len(frame) or any(width != frame.shape[1] for width in widths):
            raise RuntimeError(
                f"Mplus 数据导出结构核验失败：实际 {len(lines)} 行，期望 {len(frame)} 行；"
                f"每行应有 {frame.shape[1]} 列。"
            )
        restored = np.asarray(
            [[float(token) for token in line.decode("ascii").split()] for line in lines],
            dtype=float,
        )
        if restored.shape != expected.shape or not np.array_equal(restored, expected):
            raise RuntimeError("Mplus 数据导出回读值与转换前数据不一致，已停止运行以防变量错位。")
        if not np.array_equal(restored[:, 0].astype(int), np.arange(1, len(frame) + 1)):
            raise RuntimeError("Mplus 数据中的 ROWID 不连续，已停止运行以防个案错配。")
        result = {
            "状态": "通过",
            "编码": "ASCII（纯数字）",
            "行数": int(restored.shape[0]),
            "列数": int(restored.shape[1]),
            "每行列数一致": True,
            "ROWID连续": True,
            "数值回读一致": True,
        }
    assert result is not None
    return result


def _find_sentinel_values(series: pd.Series) -> list[float]:
    if not pd.api.types.is_numeric_dtype(series):
        return []
    vals = set(pd.to_numeric(series, errors="coerce").dropna().unique().tolist())
    return sorted(float(v) for v in vals.intersection(SENTINEL_CANDIDATES))


def prepare_lpa_project(
    input_path: str | Path,
    indicators: list[str],
    output_dir: str | Path,
    user_id: str | None = None,
    missing_codes: list[float] | None = None,
    standardize: bool = False,
    allow_low_cardinality: bool = False,
    classes: list[int] | None = None,
    text_columns: list[str] | None = None,
) -> PreparedLPAData:
    src = Path(input_path).expanduser().resolve()
    if not src.exists():
        raise FileNotFoundError(f"找不到数据文件：{src}")
    if len(indicators) < 2:
        raise ValueError("LPA 标准模块至少需要 2 个剖面指标。")
    if len(set(indicators)) != len(indicators):
        raise ValueError("剖面指标存在重复项，请确认变量清单。")

    requested_classes = classes or [1, 2, 3, 4, 5]
    if requested_classes != sorted(set(requested_classes)):
        raise ValueError("类别范围必须按升序填写且不能重复。")
    if requested_classes[0] != 1 or requested_classes != list(range(1, requested_classes[-1] + 1)):
        raise ValueError("标准 LPA 类别比较必须从 1 类开始并保持连续，例如 1,2,3,4,5。")
    if requested_classes[-1] > 10:
        raise ValueError("标准模式单次最多比较到 10 类；更大范围请进入专家模式并说明研究依据。")

    project = Path(output_dir).expanduser().resolve()
    if project.exists() and any(project.iterdir()):
        raise FileExistsError(f"输出目录已存在且非空：{project}。为避免混入旧结果，请使用新的分析目录。")
    df, source_meta = load_dataframe(
        src,
        text_columns=text_columns,
        preserve_text_columns=[user_id] if user_id else None,
    )

    missing_vars = [v for v in indicators if v not in df.columns]
    if missing_vars:
        raise ValueError(f"指定的剖面指标不存在：{missing_vars}")
    if user_id and user_id not in df.columns:
        raise ValueError(f"指定的 ID 变量不存在：{user_id}")

    missing_codes = missing_codes or []
    warnings: list[str] = []
    sentinel_findings: dict[str, list[float]] = {}

    work = df.copy()
    for code in missing_codes:
        work.loc[:, indicators] = work.loc[:, indicators].replace(code, np.nan)

    for var in indicators:
        original_series = df[var]
        sentinels = _find_sentinel_values(original_series)
        undeclared = [x for x in sentinels if x not in missing_codes]
        if undeclared:
            sentinel_findings[var] = undeclared
        if not pd.api.types.is_numeric_dtype(work[var]):
            coerced = pd.to_numeric(work[var], errors="coerce")
            newly_missing = int(coerced.isna().sum() - work[var].isna().sum())
            if newly_missing > 0:
                raise ValueError(f"剖面指标“{var}”包含无法转换为数值的内容，标准 LPA 模块已停止。")
            work[var] = coerced
        nonfinite = work[var].notna() & ~np.isfinite(pd.to_numeric(work[var], errors="coerce"))
        if nonfinite.any():
            raise ValueError(f"剖面指标“{var}”含 {int(nonfinite.sum())} 个无穷值，Mplus 无法可靠读取。")
        n_unique = int(work[var].nunique(dropna=True))
        if n_unique <= 1:
            raise ValueError(f"剖面指标“{var}”没有有效变异，不能进入 LPA。")
        if n_unique <= 10:
            if not allow_low_cardinality:
                raise ValueError(
                    f"剖面指标“{var}”只有 {n_unique} 个非缺失唯一值，可能是等级/离散变量。"
                    "标准 LPA 已停止；只有在研究者明确确认按连续变量处理后，才可使用 --confirm-low-cardinality。"
                )
            warnings.append(f"研究者已确认将低基数指标“{var}”（{n_unique} 个唯一值）按连续变量处理。")

    if sentinel_findings and not missing_codes:
        details = "；".join(f"{k}: {v}" for k, v in sentinel_findings.items())
        raise ValueError(
            "发现疑似特殊缺失值编码，但尚未声明缺失码。为防止把 -999/999 等当成真实数据，标准模式停止。"
            f" 疑似位置：{details}"
        )
    elif sentinel_findings:
        warnings.append(f"仍发现未声明的疑似缺失码：{sentinel_findings}。请确认这些确为有效观测值。")

    if user_id:
        if work[user_id].isna().any():
            warnings.append(f"用户 ID 变量“{user_id}”存在缺失；内部仍使用 ROWID 保证一一对应。")
        if work[user_id].duplicated().any():
            warnings.append(f"用户 ID 变量“{user_id}”存在重复；内部仍使用 ROWID，报告会提示该问题。")

    # 永远生成数值 ROWID，避免字符串 ID 与 Mplus 数据格式冲突。
    analysis = pd.DataFrame({"ROWID": np.arange(1, len(work) + 1, dtype=int)})
    mapping_rows: list[dict[str, Any]] = []
    means_sds: dict[str, dict[str, float]] = {}

    for idx, var in enumerate(indicators, start=1):
        internal = f"V{idx:06d}"  # 7 chars, safely under 8-char historical limit.
        s = pd.to_numeric(work[var], errors="coerce").astype(float)
        if standardize:
            mean = float(s.mean(skipna=True))
            sd = float(s.std(skipna=True, ddof=1))
            if not math.isfinite(sd) or sd == 0:
                raise ValueError(f"剖面指标“{var}”无法标准化：标准差为 0 或无效。")
            s = (s - mean) / sd
            means_sds[var] = {"均值": mean, "样本标准差": sd}
        analysis[internal] = s
        mapping_rows.append({
            "角色": "剖面指标",
            "原变量名": var,
            "Mplus内部变量名": internal,
            "原始数据类型": str(df[var].dtype),
            "非缺失数": int(s.notna().sum()),
            "缺失数": int(s.isna().sum()),
            "唯一值数": int(s.nunique(dropna=True)),
            "变量标准化": "是" if standardize else "否",
        })

    configured_temp = os.getenv("MPLUSFLOW_TEMP")
    if configured_temp:
        temp_parent = ensure_dir(Path(configured_temp).expanduser().resolve())
        runtime_dir = Path(tempfile.mkdtemp(prefix="mplusflow-run-", dir=temp_parent))
    else:
        runtime_dir = Path(tempfile.mkdtemp(prefix="mplusflow-run-"))
    if not str(runtime_dir).isascii():
        shutil.rmtree(runtime_dir, ignore_errors=True)
        raise RuntimeError(
            "Mplus 临时执行路径包含非 ASCII 字符。请把环境变量 MPLUSFLOW_TEMP 指向一个可写的纯英文路径后重试。"
        )

    dirs = {
        "原始": ensure_dir(project / "00_原始数据"),
        "检查": ensure_dir(project / "01_数据检查"),
        "分析数据": ensure_dir(project / "02_分析数据"),
        "代码": ensure_dir(project / "03_Mplus代码"),
        "原始结果": ensure_dir(project / "04_Mplus原始结果"),
        "结果": ensure_dir(project / "05_分析结果"),
        "报告": ensure_dir(project / "06_分析报告"),
        "参考": ensure_dir(project / "07_参考依据"),
        "内部": ensure_dir(project / ".mplus_runtime"),
    }
    original_copy = safe_copy(src, dirs["原始"])
    variable_map = pd.DataFrame(mapping_rows)
    variable_map_file = dirs["检查"] / "变量对应表.xlsx"
    variable_map_csv = dirs["检查"] / "变量对应表.csv"
    variable_map.to_excel(variable_map_file, index=False)
    variable_map.to_csv(variable_map_csv, index=False, encoding="utf-8-sig")

    # Mplus 预计会排除所有剖面指标均缺失的个案。
    all_missing_mask = analysis.drop(columns=["ROWID"]).isna().all(axis=1)
    expected_mplus_n = int((~all_missing_mask).sum())
    if int(all_missing_mask.sum()) > 0:
        warnings.append(f"有 {int(all_missing_mask.sum())} 个个案所有剖面指标均缺失，Mplus 可能不会将其纳入模型估计。")

    data_file = dirs["分析数据"] / "Mplus分析数据.dat"
    # 内部运行副本使用 ASCII 文件名，内容只含 ASCII 数字。
    runtime_data = runtime_dir / "data.dat"
    export_check = write_verified_mplus_data(analysis, [data_file, runtime_data])

    size_advisory = sample_size_advisory("lpa", expected_mplus_n)
    spec = {
        "分析类型": "LPA",
        "模板ID": "LPA-BASE-EV0",
        "数据文件": str(src),
        "数据文件SHA256": sha256_file(src),
        "用户ID变量": user_id,
        "剖面指标": indicators,
        "内部剖面指标": variable_map["Mplus内部变量名"].tolist(),
        "内部ID变量": "ROWID",
        "类别范围": requested_classes,
        "变量标准化": bool(standardize),
        "低基数指标连续化确认": bool(allow_low_cardinality),
        "标准化参数": means_sds,
        "用户声明缺失码": missing_codes,
        "内部缺失码": INTERNAL_MISSING,
        "原始样本数": int(len(df)),
        "预计Mplus有效样本数": expected_mplus_n,
        "源数据元信息": source_meta,
        "Mplus数据转换核验": export_check,
        "数据警告": warnings,
        "样本量提示": size_advisory,
    }
    spec_file = dirs["分析数据"] / "分析设计清单.json"
    write_json(spec_file, spec)

    audit_lines = [
        "# 数据质量检查报告",
        "",
        f"- 原始数据：`{src.name}`",
        f"- 原始样本数：{len(df)}",
        f"- 剖面指标数：{len(indicators)}",
        f"- 预计 Mplus 有效样本数：{expected_mplus_n}",
        f"- 是否对指标做 Z 标准化：{'是' if standardize else '否'}",
        f"- 内部缺失码：{int(INTERNAL_MISSING)}",
        f"- 数据转换核验：{export_check['状态']}；{export_check['行数']} 行 × {export_check['列数']} 列；{export_check['编码']}",
        "",
        "## 剖面指标",
        "",
    ]
    if source_meta.get("文本编码"):
        audit_lines.insert(4, f"- 识别到的源文本编码：{source_meta['文本编码']}")
    for _, row in variable_map.iterrows():
        audit_lines.append(
            f"- {row['原变量名']} → {row['Mplus内部变量名']}；非缺失 {row['非缺失数']}；缺失 {row['缺失数']}；唯一值 {row['唯一值数']}"
        )
    audit_lines.extend(["", "## 警告", ""])
    if warnings:
        audit_lines.extend(f"- {w}" for w in warnings)
    else:
        audit_lines.append("- 未发现阻止标准 LPA 运行的数据问题。")
    audit_lines.extend(["", *advisory_markdown(size_advisory)])
    audit_file = dirs["检查"] / "数据质量检查报告.md"
    audit_file.write_text("\n".join(audit_lines), encoding="utf-8")

    # 用户ID与ROWID映射单独保存，避免字符串ID进入Mplus。
    id_map = pd.DataFrame({"ROWID": analysis["ROWID"]})
    if user_id:
        id_map[user_id] = df[user_id].map(lambda value: None if pd.isna(value) else str(value))
    id_map.to_excel(dirs["检查"] / "个案ID对应表.xlsx", index=False)
    id_map.to_csv(dirs["检查"] / "个案ID对应表.csv", index=False, encoding="utf-8-sig")

    return PreparedLPAData(
        project_dir=project,
        runtime_dir=runtime_dir,
        audit_runtime_dir=dirs["内部"],
        data_file=data_file,
        original_copy=original_copy,
        spec_file=spec_file,
        audit_file=audit_file,
        variable_map_file=variable_map_file,
        variable_map_csv=variable_map_csv,
        spec=spec,
        variable_map=variable_map,
        id_map=id_map,
        analysis_df=analysis,
        expected_mplus_n=expected_mplus_n,
    )
